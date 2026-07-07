"""Password change script for remote Windows servers.

Changes the local Administrator password on each server in the
server list.  The same new password is applied to every server.
Generates a password change report in Excel format.
"""

import argparse
import logging
import os
import sys
from datetime import datetime
from getpass import getpass

import requests.exceptions
import winrm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from winrm.exceptions import AuthenticationError, WinRMTransportError

from crypto import read_encrypted_file

logger = logging.getLogger(__name__)


def load_servers(servers_path: str) -> list[str]:
    """Read server addresses from a text file.

    One hostname or IP per line. Blank lines and lines starting with ``#``
    are ignored.

    Args:
        servers_path: Path to the servers text file.

    Returns:
        List of server address strings.

    Raises:
        FileNotFoundError: If the file does not exist.
    """
    servers = []
    with open(servers_path, "r") as f:
        for line in f:
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            servers.append(stripped)
    return servers


def create_session(
    server: str,
    usernames: list[str],
    passwords: list[str],
    try_creds: bool = True,
) -> tuple:
    """Create a WinRM session to a server with credential fallback.

    Tries every combination of usernames and passwords, first with CredSSP
    transport then NTLM as a fallback.  On ``AuthenticationError`` the
    inner transport loop is skipped for that password (wrong creds); all
    other errors continue searching.

    Args:
        server: Hostname or IP of the target server.
        usernames: Ordered list of usernames to try (e.g.
            ``['administrator', 'admin']``).
        passwords: List of candidate passwords.
        try_creds: If ``False`` only the first username and first password
            are attempted (no fallback).  Default ``True``.

    Returns:
        ``(session, username)`` on success, ``(None, error_message)`` on
        failure.
    """
    transports = ["credssp", "ntlm"]

    if not try_creds:
        usernames = usernames[:1]
        passwords = passwords[:1]

    errors: list[str] = []
    for user in usernames:
        for pwd in passwords:
            for transport in transports:
                try:
                    session = winrm.Session(
                        server,
                        auth=(user, pwd),
                        transport=transport,
                        server_cert_validation="ignore",
                    )
                    rs = session.run_ps("Write-Output 'ping'")
                    if rs.status_code == 0:
                        logger.info(
                            "Connected to %s as %s via %s", server, user, transport
                        )
                        return session, user
                except AuthenticationError:
                    errors.append(f"Authentication failed: {user} via {transport}")
                    break
                except WinRMTransportError as e:
                    errors.append(
                        f"WinRMTransportError for {user} via {transport}: {e}"
                    )
                except requests.exceptions.ConnectionError as e:
                    errors.append(f"ConnectionError for {server}: {e}")
                except requests.exceptions.Timeout as e:
                    errors.append(f"Timeout for {server}: {e}")
                except Exception as e:
                    errors.append(
                        f"Unexpected error for {user} via {transport}: {e}"
                    )

    return None, "; ".join(errors) or "All credential combinations failed"


def run_ps(session, script: str) -> tuple:
    """Run a PowerShell script on a WinRM session.

    Args:
        session: A ``winrm.Session`` instance.
        script: PowerShell script text to execute.

    Returns:
        ``(True, stdout_string)`` on success,
        ``(False, {"error": ..., "status_code": ...})`` on failure.
    """
    try:
        rs = session.run_ps(script)
        if rs.status_code == 0:
            return True, rs.std_out.decode("utf-8", errors="replace").strip()
        else:
            err_text = rs.std_err.decode("utf-8", errors="replace").strip()
            return False, {"error": err_text, "status_code": rs.status_code}
    except WinRMTransportError as e:
        return False, {
            "error": f"WinRM transport error: {e}",
            "status_code": e.code if hasattr(e, "code") else 0,
        }
    except AuthenticationError as e:
        return False, {
            "error": f"Authentication error: {e}",
            "status_code": 401,
        }
    except requests.exceptions.Timeout as e:
        return False, {
            "error": f"Operation timed out: {e}",
            "status_code": 0,
        }
    except requests.exceptions.ConnectionError as e:
        return False, {
            "error": f"Connection lost: {e}",
            "status_code": 0,
        }
    except Exception as e:
        return False, {
            "error": f"Unexpected error: {e}",
            "status_code": 0,
        }


def close_session(session) -> None:
    """Release resources held by a WinRM session.

    Safe to call multiple times or with ``None``.

    Args:
        session: A ``winrm.Session`` instance or ``None``.
    """
    if session is not None:
        try:
            session.protocol.transport.close_session()
        except Exception:
            pass


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(
    start_color="2F5496", end_color="2F5496", fill_type="solid"
)
HEADER_ALIGNMENT = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_workbook() -> Workbook:
    """Create a new Excel workbook with the default sheet removed.

    Returns:
        An empty ``Workbook`` instance ready for worksheets to be added.
    """
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _add_header_style(ws, headers: list[str]) -> None:
    """Write formatted header row and configure auto-filter and freeze.

    Writes headers to row 1, applies font, fill, alignment, and border
    styling, adds an auto-filter over the header range, and freezes the
    top row for scrolling.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        headers: Column header strings to write in order.
    """
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.freeze_panes = "A2"


def _auto_width(ws) -> None:
    """Adjust column widths based on content length, capped at 60.

    Iterates over every column in the worksheet and sets the column
    width to the length of the longest value (header or cell) plus a
    padding of 2, with a maximum width of 60 characters.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
    """
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except (ValueError, TypeError):
                pass
        adjusted = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = adjusted


def write_password_status(ws, data: list[dict]) -> None:
    """Write the Password Change Status worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``old_connection_status``,
            ``new_password_set``, ``error``.
    """
    headers = ["Server", "Old Connection Status", "New Password Set", "Error"]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(
            row=row_idx,
            column=2,
            value=entry.get("old_connection_status", ""),
        )
        ws.cell(
            row=row_idx,
            column=3,
            value=entry.get("new_password_set", False),
        )
        ws.cell(row=row_idx, column=4, value=entry.get("error", ""))
    _auto_width(ws)


def write_errors(ws, data: list[dict]) -> None:
    """Write the Errors worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``operation``,
            ``error``, ``timestamp``.
    """
    headers = ["Server", "Operation", "Error", "Timestamp"]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("operation", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("error", ""))
        ws.cell(row=row_idx, column=4, value=entry.get("timestamp", ""))
    _auto_width(ws)


def save_report(workbook, path: str) -> None:
    """Save a workbook to disk.

    Args:
        workbook: An openpyxl ``Workbook`` instance to save.
        path: Destination file path.
    """
    workbook.save(path)


def _setup_file_logging(log_path: str) -> None:
    log_dir = os.path.dirname(log_path)
    if log_dir:
        os.makedirs(log_dir, exist_ok=True)
    handler = logging.FileHandler(log_path, encoding="utf-8", mode="w")
    handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    )
    logging.getLogger().addHandler(handler)
    logging.getLogger().setLevel(logging.INFO)
    logging.getLogger("openpyxl").setLevel(logging.WARNING)
    logging.getLogger("urllib3").setLevel(logging.WARNING)


_CHANGE_PASSWORD_SCRIPT = """
$username = 'Administrator'
try {
    $newPassword = ConvertTo-SecureString -String '{password}' -AsPlainText -Force
    $user = Get-LocalUser -Name $username -ErrorAction Stop
    $user | Set-LocalUser -Password $newPassword -ErrorAction Stop
    "SUCCESS"
} catch {
    "FAILED: $($_.Exception.Message)"
}
"""


def prompt_new_password() -> str:
    """Prompt the user for a new Administrator password with confirmation.

    Returns:
        The confirmed new password string.

    Raises:
        SystemExit: If the passwords do not match.
    """
    pwd = getpass("Enter new Administrator password: ")
    confirm = getpass("Confirm new Administrator password: ")
    if pwd != confirm:
        print("Error: Passwords do not match.", file=sys.stderr)
        sys.exit(1)
    return pwd


def change_password(session, new_password: str) -> tuple:
    """Change the local Administrator password on a remote server.

    Args:
        session: A ``winrm.Session`` instance.
        new_password: The new password to set on the Administrator account.

    Returns:
        ``(True, "")`` on success,
        ``(False, error_message)`` on failure.
    """
    escaped = new_password.replace("'", "''")
    script = _CHANGE_PASSWORD_SCRIPT.replace("{password}", escaped)
    success, result = run_ps(session, script)
    if not success:
        return False, result.get("error", "Unknown error")

    stdout = result.strip() if isinstance(result, str) else ""
    if stdout == "SUCCESS":
        return True, ""
    elif stdout.startswith("FAILED:"):
        return False, stdout[7:].strip()
    else:
        return False, stdout or "Unknown result"


def main():
    """Entry point for the password change script."""
    parser = argparse.ArgumentParser(
        description="Change local Administrator password on remote servers"
    )
    parser.add_argument(
        "--servers",
        default=os.path.join("config", "servers.txt"),
        help="Path to server list (default: config/servers.txt)",
    )
    parser.add_argument(
        "--key",
        default=os.path.join("config", "secret.key"),
        help="Path to encryption key (default: config/secret.key)",
    )
    parser.add_argument(
        "--passwords",
        default=os.path.join("config", "passwords.enc"),
        help="Path to encrypted passwords (default: config/passwords.enc)",
    )
    parser.add_argument(
        "--output",
        default="output",
        help="Output directory for reports (default: output/)",
    )
    args = parser.parse_args()

    os.makedirs(args.output, exist_ok=True)
    ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
    _setup_file_logging(os.path.join(args.output, f"Password_Change_{ts_file}.log"))

    servers = load_servers(args.servers)
    if not servers:
        print("No servers found in server list.", file=sys.stderr)
        logger.error("No servers found in server list.")
        sys.exit(1)

    if not os.path.exists(args.key):
        print(
            f"Error: Encryption key not found at {args.key}",
            file=sys.stderr,
        )
        logger.error("Encryption key not found at %s", args.key)
        sys.exit(1)

    if not os.path.exists(args.passwords):
        print(
            f"Error: Encrypted passwords not found at {args.passwords}",
            file=sys.stderr,
        )
        logger.error("Encrypted passwords not found at %s", args.passwords)
        sys.exit(1)

    passwords = read_encrypted_file(args.key, args.passwords)
    usernames = ["administrator", "admin"]
    new_password = prompt_new_password()
    print()

    password_status_rows: list[dict] = []
    error_rows: list[dict] = []
    success_count = 0
    failed_count = 0

    for server in servers:
        logger.info("[%s] Connecting...", server)
        sys.stdout.write(f"[{server}] Connecting... ")
        sys.stdout.flush()
        session, result = create_session(server, usernames, passwords)

        if session is not None:
            username = result
            old_status = f"Connected as {username}"
            logger.info("[%s] %s", server, old_status)
            sys.stdout.write(f"{old_status}... ")
            sys.stdout.flush()

            pwd_success, pwd_msg = change_password(session, new_password)
            close_session(session)

            if pwd_success:
                logger.info("[%s] password changed SUCCESS", server)
                print("password changed SUCCESS")
                success_count += 1
            else:
                logger.error("[%s] password changed FAILED - %s", server, pwd_msg)
                print(f"password changed FAILED - {pwd_msg}")
                failed_count += 1

            password_status_rows.append({
                "server": server,
                "old_connection_status": old_status,
                "new_password_set": pwd_success,
                "error": pwd_msg if not pwd_success else "",
            })
            if not pwd_success:
                error_rows.append({
                    "server": server,
                    "operation": "change_password",
                    "error": pwd_msg,
                    "timestamp": datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                })
        else:
            logger.error("[%s] Connection FAILED - %s", server, result)
            print(f"Connection FAILED - {result}")
            failed_count += 1
            password_status_rows.append({
                "server": server,
                "old_connection_status": "Connection FAILED",
                "new_password_set": False,
                "error": result,
            })
            error_rows.append({
                "server": server,
                "operation": "connect",
                "error": result,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })

    print("---")
    logger.info(
        "Password change complete. %d succeeded, %d failed.",
        success_count, failed_count,
    )
    print(
        f"Password change complete. "
        f"{success_count} succeeded, {failed_count} failed."
    )

    filename = f"Password_Change_{ts_file}.xlsx"
    output_path = os.path.join(args.output, filename)

    wb = create_workbook()
    ws_status = wb.create_sheet("Password Change Status")
    write_password_status(ws_status, password_status_rows)
    if error_rows:
        ws_errors = wb.create_sheet("Errors")
        write_errors(ws_errors, error_rows)
    save_report(wb, output_path)
    report_path = os.path.abspath(output_path)
    logger.info("Report saved to %s", report_path)
    print(f"Report saved to {report_path}")


if __name__ == "__main__":
    main()
