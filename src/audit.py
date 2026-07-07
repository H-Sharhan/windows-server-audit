"""Read-only audit orchestrator.

Connects to each server in the server list, collects local users, groups,
and scheduled tasks, and generates a formatted Excel report.
"""

import argparse
import json
import logging
import os
import sys
from datetime import datetime

import requests.exceptions
import winrm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from winrm.exceptions import AuthenticationError, WinRMTransportError

from crypto import read_encrypted_file

logger = logging.getLogger(__name__)

_BUILTIN_ACCOUNTS = {"administrator", "guest", "defaultaccount"}

_USERS_SCRIPT = """
$privilegedGroups = @(
    'Administrators',
    'Backup Operators',
    'Power Users',
    'Remote Desktop Users',
    'Remote Management Users',
    'Hyper-V Administrators'
)
$results = @()
foreach ($user in Get-LocalUser) {
    $memberships = @()
    foreach ($group in $privilegedGroups) {
        $members = net localgroup $group 2>$null
        if ($members -contains $user.Name) {
            $memberships += $group
        }
    }
    $results += [PSCustomObject]@{
        Name             = $user.Name
        FullName         = $user.FullName
        Enabled          = $user.Enabled
        PrivilegedGroups = $memberships -join ', '
        SID              = $user.SID.Value
        PrincipalSource  = $user.PrincipalSource
        Description      = $user.Description
    }
}
$results | ConvertTo-Json
"""

_PRIORITIZED_GROUPS = frozenset({
    "Administrators",
    "Remote Desktop Users",
    "Backup Operators",
    "Power Users",
    "Remote Management Users",
    "Hyper-V Administrators",
})

_WELL_KNOWN_NAME_PREFIXES = (
    "NT AUTHORITY\\",
    "BUILTIN\\",
    "APPLICATION PACKAGE AUTHORITY\\",
)

_GROUPS_SCRIPT = """
$groups = Get-LocalGroup | Select-Object Name, Description, @{N='SID';E={$_.SID.Value}}
$results = @()
foreach ($group in $groups) {
    $members = @()
    try {
        $groupMembers = Get-LocalGroupMember -Group $group.Name -ErrorAction Stop
        foreach ($m in $groupMembers) {
            $enabled = $null
            if ($m.ObjectClass -eq 'User') {
                try {
                    $adObj = [ADSI]"WinNT://$env:COMPUTERNAME/$($m.Name.Split('\')[-1])"
                    $enabled = -not [bool]$adObj.AccountDisabled
                } catch {
                    $enabled = $null
                }
            }
            $members += [PSCustomObject]@{
                Name        = $m.Name
                SID         = $m.SID.Value
                ObjectClass = $m.ObjectClass
                PrincipalSource = $m.PrincipalSource
                Enabled     = $enabled
            }
        }
    } catch {
        $members = @()
    }
    $results += [PSCustomObject]@{
        GroupName   = $group.Name
        Description = $group.Description
        SID         = $group.SID
        Members     = $members
    }
}
$results | ConvertTo-Json -Depth 3
"""

_TASKS_SCRIPT = """
Get-ScheduledTask | ForEach-Object {
    $task = $_
    $info = Get-ScheduledTaskInfo -TaskPath $task.TaskPath -TaskName $task.TaskName -ErrorAction SilentlyContinue
    $actions = if ($task.Actions) { $task.Actions | ForEach-Object { $_.Execute + ' ' + $_.Arguments } } else { @() }
    $triggers = if ($task.Triggers) { $task.Triggers | ForEach-Object { $_.ToString() } } else { @() }
    $principal = $task.Principal
    [PSCustomObject]@{
        TaskName     = $task.TaskName
        TaskPath     = $task.TaskPath
        RunAsUser    = if ($principal -and $principal.UserId) { $principal.UserId } else { 'SYSTEM' }
        Triggers     = ($triggers | Where-Object { $_ }) -join '; '
        Actions      = ($actions | Where-Object { $_ }) -join '; '
        State        = $task.State
        LastRunTime  = if ($info -and $info.LastRunTime) { $info.LastRunTime.ToString('yyyy-MM-dd HH:mm:ss') } else { '' }
        NextRunTime  = if ($info -and $info.NextRunTime) { $info.NextRunTime.ToString('yyyy-MM-dd HH:mm:ss') } else { '' }
    }
} | ConvertTo-Json -Depth 3
"""


def load_servers(servers_path: str) -> list[str]:
    """Read server addresses from a text file.

    One hostname or IP per line. Blank lines and lines starting with ``#``
    are ignored.

    Args:
        servers_path: Path to the servers text file.

    Returns:
        List of server address strings.

    Raises:
        FileNotFoundError: If the file does not exist.
    """
    servers = []
    with open(servers_path, "r") as f:
        for line in f:
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            servers.append(stripped)
    return servers


def create_session(
    server: str,
    usernames: list[str],
    passwords: list[str],
    try_creds: bool = True,
) -> tuple:
    """Create a WinRM session to a server with credential fallback.

    Tries every combination of usernames and passwords, first with CredSSP
    transport then NTLM as a fallback.  On ``AuthenticationError`` the
    inner transport loop is skipped for that password (wrong creds); all
    other errors continue searching.

    Args:
        server: Hostname or IP of the target server.
        usernames: Ordered list of usernames to try (e.g.
            ``['administrator', 'admin']``).
        passwords: List of candidate passwords.
        try_creds: If ``False`` only the first username and first password
            are attempted (no fallback).  Default ``True``.

    Returns:
        ``(session, username)`` on success, ``(None, error_message)`` on
        failure.
    """
    transports = ["credssp", "ntlm"]

    if not try_creds:
        usernames = usernames[:1]
        passwords = passwords[:1]

    errors: list[str] = []
    for user in usernames:
        for pwd in passwords:
            for transport in transports:
                try:
                    session = winrm.Session(
                        server,
                        auth=(user, pwd),
                        transport=transport,
                        server_cert_validation="ignore",
                    )
                    rs = session.run_ps("Write-Output 'ping'")
                    if rs.status_code == 0:
                        logger.info(
                            "Connected to %s as %s via %s", server, user, transport
                        )
                        return session, user
                except AuthenticationError:
                    errors.append(f"Authentication failed: {user} via {transport}")
                    break
                except WinRMTransportError as e:
                    errors.append(
                        f"WinRMTransportError for {user} via {transport}: {e}"
                    )
                except requests.exceptions.ConnectionError as e:
                    errors.append(f"ConnectionError for {server}: {e}")
                except requests.exceptions.Timeout as e:
                    errors.append(f"Timeout for {server}: {e}")
                except Exception as e:
                    errors.append(
                        f"Unexpected error for {user} via {transport}: {e}"
                    )

    return None, "; ".join(errors) or "All credential combinations failed"


def run_ps(session, script: str) -> tuple:
    """Run a PowerShell script on a WinRM session.

    Args:
        session: A ``winrm.Session`` instance.
        script: PowerShell script text to execute.

    Returns:
        ``(True, stdout_string)`` on success,
        ``(False, {"error": ..., "status_code": ...})`` on failure.
    """
    try:
        rs = session.run_ps(script)
        if rs.status_code == 0:
            return True, rs.std_out.decode("utf-8", errors="replace").strip()
        else:
            err_text = rs.std_err.decode("utf-8", errors="replace").strip()
            return False, {"error": err_text, "status_code": rs.status_code}
    except WinRMTransportError as e:
        return False, {
            "error": f"WinRM transport error: {e}",
            "status_code": e.code if hasattr(e, "code") else 0,
        }
    except AuthenticationError as e:
        return False, {
            "error": f"Authentication error: {e}",
            "status_code": 401,
        }
    except requests.exceptions.Timeout as e:
        return False, {
            "error": f"Operation timed out: {e}",
            "status_code": 0,
        }
    except requests.exceptions.ConnectionError as e:
        return False, {
            "error": f"Connection lost: {e}",
            "status_code": 0,
        }
    except Exception as e:
        return False, {
            "error": f"Unexpected error: {e}",
            "status_code": 0,
        }


def close_session(session) -> None:
    """Release resources held by a WinRM session.

    Safe to call multiple times or with ``None``.

    Args:
        session: A ``winrm.Session`` instance or ``None``.
    """
    if session is not None:
        try:
            session.protocol.transport.close_session()
        except Exception:
            pass


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(
    start_color="2F5496", end_color="2F5496", fill_type="solid"
)
HEADER_ALIGNMENT = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_workbook() -> Workbook:
    """Create a new Excel workbook with the default sheet removed.

    Returns:
        An empty ``Workbook`` instance ready for worksheets to be added.
    """
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _add_header_style(ws, headers: list[str]) -> None:
    """Write formatted header row and configure auto-filter and freeze.

    Writes headers to row 1, applies font, fill, alignment, and border
    styling, adds an auto-filter over the header range, and freezes the
    top row for scrolling.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        headers: Column header strings to write in order.
    """
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.freeze_panes = "A2"


def _auto_width(ws) -> None:
    """Adjust column widths based on content length, capped at 60.

    Iterates over every column in the worksheet and sets the column
    width to the length of the longest value (header or cell) plus a
    padding of 2, with a maximum width of 60 characters.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
    """
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except (ValueError, TypeError):
                pass
        adjusted = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = adjusted


def write_summary(ws, data: list[dict]) -> None:
    """Write the Summary worksheet with per-server overview counts.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``total_users``,
            ``admins``, ``privileged``, ``disabled``, ``builtin``,
            ``groups``, ``tasks``, ``login_status``.
    """
    headers = [
        "Server",
        "Total Users",
        "Admins",
        "Privileged",
        "Disabled",
        "Built-in",
        "Groups",
        "Tasks",
        "Login Status",
    ]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("total_users", 0))
        ws.cell(row=row_idx, column=3, value=entry.get("admins", 0))
        ws.cell(row=row_idx, column=4, value=entry.get("privileged", 0))
        ws.cell(row=row_idx, column=5, value=entry.get("disabled", 0))
        ws.cell(row=row_idx, column=6, value=entry.get("builtin", 0))
        ws.cell(row=row_idx, column=7, value=entry.get("groups", 0))
        ws.cell(row=row_idx, column=8, value=entry.get("tasks", 0))
        ws.cell(row=row_idx, column=9, value=entry.get("login_status", ""))
    _auto_width(ws)


def write_login_status(ws, data: list[dict]) -> None:
    """Write the Server Login Status worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``success``,
            ``username_used``, ``error``.
    """
    headers = ["Server", "Success", "Username Used", "Error"]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("success", False))
        ws.cell(
            row=row_idx, column=3, value=entry.get("username_used", "")
        )
        ws.cell(row=row_idx, column=4, value=entry.get("error", ""))
    _auto_width(ws)


def write_users(ws, data: list[dict]) -> None:
    """Write the Local Users and Privileges worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``name``, ``fullname``,
            ``enabled``, ``privilege_level``, ``privileged_groups``,
            ``principal_source``, ``is_builtin``.
    """
    headers = [
        "Server",
        "Name",
        "Full Name",
        "Enabled",
        "Privilege Level",
        "Privileged Groups",
        "Principal Source",
        "Is Built-in",
    ]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("name", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("fullname", ""))
        ws.cell(row=row_idx, column=4, value=entry.get("enabled", False))
        ws.cell(
            row=row_idx,
            column=5,
            value=entry.get("privilege_level", ""),
        )
        ws.cell(
            row=row_idx,
            column=6,
            value=entry.get("privileged_groups", ""),
        )
        ws.cell(
            row=row_idx,
            column=7,
            value=entry.get("principal_source", ""),
        )
        ws.cell(row=row_idx, column=8, value=entry.get("is_builtin", False))
    _auto_width(ws)


def write_groups(ws, data: list[dict]) -> None:
    """Write the Local Groups and Members worksheet.

    Each dict represents one group-member row (data must be flattened
    before calling this function).

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``group_name``,
            ``member_name``, ``account_type``, ``enabled``,
            ``principal_source``.
    """
    headers = [
        "Server",
        "Group Name",
        "Member Name",
        "Account Type",
        "Enabled",
        "Principal Source",
    ]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("group_name", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("member_name", ""))
        ws.cell(
            row=row_idx, column=4, value=entry.get("account_type", "")
        )
        ws.cell(row=row_idx, column=5, value=entry.get("enabled"))
        ws.cell(
            row=row_idx,
            column=6,
            value=entry.get("principal_source", ""),
        )
    _auto_width(ws)


def write_tasks(ws, data: list[dict]) -> None:
    """Write the Scheduled Tasks worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``task_name``,
            ``task_path``, ``run_as_user``, ``triggers``, ``actions``,
            ``state``, ``last_run_time``, ``next_run_time``,
            ``classification``.
    """
    headers = [
        "Server",
        "Task Name",
        "Path",
        "Run As User",
        "Triggers",
        "Actions",
        "State",
        "Last Run",
        "Next Run",
        "Classification",
    ]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("task_name", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("task_path", ""))
        ws.cell(
            row=row_idx, column=4, value=entry.get("run_as_user", "")
        )
        ws.cell(row=row_idx, column=5, value=entry.get("triggers", ""))
        ws.cell(row=row_idx, column=6, value=entry.get("actions", ""))
        ws.cell(row=row_idx, column=7, value=entry.get("state", ""))
        ws.cell(
            row=row_idx, column=8, value=entry.get("last_run_time", "")
        )
        ws.cell(
            row=row_idx, column=9, value=entry.get("next_run_time", "")
        )
        ws.cell(
            row=row_idx,
            column=10,
            value=entry.get("classification", ""),
        )
    _auto_width(ws)


def write_errors(ws, data: list[dict]) -> None:
    """Write the Errors worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``operation``,
            ``error``, ``timestamp``.
    """
    headers = ["Server", "Operation", "Error", "Timestamp"]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("operation", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("error", ""))
        ws.cell(row=row_idx, column=4, value=entry.get("timestamp", ""))
    _auto_width(ws)


def save_report(workbook, path: str) -> None:
    """Save a workbook to disk.

    Args:
        workbook: An openpyxl ``Workbook`` instance to save.
        path: Destination file path.
    """
    workbook.save(path)


def _classify_privilege(
    name: str,
    enabled: bool,
    privileged_groups: list[str],
    principal_source: str,
) -> str:
    """Classify a user's privilege level.

    Priority order: Administrators → Local Administrator.
    Backup Operators / Power Users / Hyper-V Administrators → Privileged User.
    Disabled account → Disabled User.
    Well-known built-in local account → Built-in/System Account.
    Everything else → Standard User.

    Args:
        name: The user's account name.
        enabled: Whether the account is enabled.
        privileged_groups: List of privileged group names the user belongs to.
        principal_source: The source of the account (e.g. 'Local').

    Returns:
        A string privilege level classification.
    """
    if "Administrators" in privileged_groups:
        return "Local Administrator"
    for group in ("Backup Operators", "Power Users", "Hyper-V Administrators"):
        if group in privileged_groups:
            return "Privileged User"
    if not enabled:
        return "Disabled User"
    if principal_source == "Local" and name.lower() in _BUILTIN_ACCOUNTS:
        return "Built-in/System Account"
    return "Standard User"


def collect_users(session) -> list[dict]:
    """Enumerate local users and their privileges on a remote server.

    Runs a PowerShell script via WinRM that collects all local user
    accounts and checks their membership in well-known privileged groups.
    Each user is then classified into a privilege level.

    Args:
        session: A ``winrm.Session`` instance obtained from
            ``create_session``.

    Returns:
        A list of dictionaries, one per user, with keys:
        ``name``, ``fullname``, ``enabled``, ``privilege_level``,
        ``privileged_groups``, ``sid``, ``principal_source``,
        ``description``, ``is_builtin``.
        Returns an empty list if the operation fails on the server.
    """
    success, result = run_ps(session, _USERS_SCRIPT)
    if not success:
        logger.error(
            "Failed to enumerate local users: %s",
            result.get("error", "Unknown error"),
        )
        return []

    if not result:
        return []

    try:
        raw_users = json.loads(result)
    except json.JSONDecodeError as e:
        logger.error("Failed to parse user enumeration JSON: %s", e)
        return []

    if isinstance(raw_users, dict):
        raw_users = [raw_users]
    elif not isinstance(raw_users, list):
        logger.warning(
            "Unexpected JSON structure from user enumeration: %s",
            type(raw_users),
        )
        return []

    users = []
    for raw in raw_users:
        if not isinstance(raw, dict):
            continue

        name = raw.get("Name", "") or ""
        fullname = raw.get("FullName", "") or ""
        enabled = bool(raw.get("Enabled", False))
        sid = raw.get("SID", "") or ""
        principal_source = raw.get("PrincipalSource", "") or ""
        description = raw.get("Description", "") or ""
        groups_str = raw.get("PrivilegedGroups", "") or ""
        privileged_groups = [
            g.strip() for g in groups_str.split(",") if g.strip()
        ]

        is_builtin = (
            principal_source == "Local"
            and name.lower() in _BUILTIN_ACCOUNTS
        )

        privilege_level = _classify_privilege(
            name, enabled, privileged_groups, principal_source
        )

        users.append({
            "name": name,
            "fullname": fullname,
            "enabled": enabled,
            "privilege_level": privilege_level,
            "privileged_groups": groups_str,
            "sid": sid,
            "principal_source": principal_source,
            "description": description,
            "is_builtin": is_builtin,
        })

    return users


def _is_well_known_name(name: str) -> bool:
    """Check if a member name indicates a well-known built-in account.

    Args:
        name: The member name (e.g. ``NT AUTHORITY\\NETWORK SERVICE``).

    Returns:
        True if the name starts with a well-known authority prefix.
    """
    return name.upper().startswith(_WELL_KNOWN_NAME_PREFIXES)


def _classify_account_type(
    object_class: str,
    principal_source: str,
    name: str,
    sid: str,
) -> str:
    """Classify a group member's account type.

    Classification follows the rules in the project specification:

    - ``User`` + ``Local`` → ``Local User``
    - ``User`` + ``ActiveDirectory`` → ``Domain User``
    - ``User`` + ``MicrosoftAccount`` → ``Microsoft Account``
    - ``Group`` + ``Local`` → ``Local Group``
    - ``Group`` + ``ActiveDirectory`` → ``Domain Group``
    - Well-known names (NT AUTHORITY\\*, BUILTIN\\*, etc.) → ``Built-in Account``
    - Everything else → ``Built-in Account``

    Args:
        object_class: ``'User'`` or ``'Group'``.
        principal_source: e.g. ``'Local'``, ``'ActiveDirectory'``,
            ``'MicrosoftAccount'``.
        name: The member's full name (e.g. ``'DOMAIN\\User'``).
        sid: The member's security identifier string.

    Returns:
        One of ``'Local User'``, ``'Domain User'``, ``'Microsoft Account'``,
        ``'Local Group'``, ``'Domain Group'``, ``'Built-in Account'``.
    """
    if _is_well_known_name(name):
        return "Built-in Account"

    if object_class == "User":
        if principal_source == "Local":
            return "Local User"
        elif principal_source == "ActiveDirectory":
            return "Domain User"
        elif principal_source == "MicrosoftAccount":
            return "Microsoft Account"
    elif object_class == "Group":
        if principal_source == "Local":
            return "Local Group"
        elif principal_source == "ActiveDirectory":
            return "Domain Group"

    return "Built-in Account"


def _is_prioritized_group(group_name: str) -> bool:
    """Check if the group is in the prioritized reporting list.

    Prioritized groups include Administrators, Remote Desktop Users,
    Backup Operators, Power Users, Remote Management Users, and
    Hyper-V Administrators.

    Args:
        group_name: The local group name.

    Returns:
        True if the group is in the prioritized list.
    """
    return group_name in _PRIORITIZED_GROUPS


def collect_groups(session) -> list[dict]:
    """Enumerate all local groups and their members on a remote server.

    Runs a PowerShell script via WinRM that collects every local group,
    their descriptions, SIDs, and members.  For each member the account
    type is classified and the enabled/disabled status is checked for
    user-type members.

    Args:
        session: A ``winrm.Session`` instance obtained from
            ``create_session``.

    Returns:
        A list of dictionaries, one per group, with keys:
        ``group_name``, ``description``, ``sid``, ``is_prioritized``,
        ``members``.

        Each member dict has keys: ``name``, ``sid``, ``object_class``,
        ``principal_source``, ``account_type``, ``enabled``.

        Returns an empty list if the operation fails on the server.
    """
    success, result = run_ps(session, _GROUPS_SCRIPT)
    if not success:
        logger.error(
            "Failed to enumerate local groups: %s",
            result.get("error", "Unknown error"),
        )
        return []

    if not result:
        return []

    try:
        raw_groups = json.loads(result)
    except json.JSONDecodeError as e:
        logger.error("Failed to parse group enumeration JSON: %s", e)
        return []

    if isinstance(raw_groups, dict):
        raw_groups = [raw_groups]
    elif not isinstance(raw_groups, list):
        logger.warning(
            "Unexpected JSON structure from group enumeration: %s",
            type(raw_groups),
        )
        return []

    groups = []
    for raw in raw_groups:
        if not isinstance(raw, dict):
            continue

        group_name = raw.get("GroupName", "") or ""
        description = raw.get("Description", "") or ""
        sid = raw.get("SID", "") or ""
        raw_members = raw.get("Members", [])

        if not isinstance(raw_members, list):
            raw_members = []

        members = []
        for m in raw_members:
            if not isinstance(m, dict):
                continue

            m_name = m.get("Name", "") or ""
            m_sid = m.get("SID", "") or ""
            m_object_class = m.get("ObjectClass", "") or ""
            m_principal_source = m.get("PrincipalSource", "") or ""
            m_enabled = m.get("Enabled")

            account_type = _classify_account_type(
                m_object_class,
                m_principal_source,
                m_name,
                m_sid,
            )

            members.append({
                "name": m_name,
                "sid": m_sid,
                "object_class": m_object_class,
                "principal_source": m_principal_source,
                "account_type": account_type,
                "enabled": m_enabled,
            })

        groups.append({
            "group_name": group_name,
            "description": description,
            "sid": sid,
            "is_prioritized": _is_prioritized_group(group_name),
            "members": members,
        })

    return groups


def classify_task(task_path: str) -> str:
    """Classify a scheduled task as system/default or user-created.

    Tasks stored under ``\\Microsoft\\`` or ``\\Windows\\`` paths are
    considered system/default tasks. All others are user-created.

    Args:
        task_path: The task path
            (e.g. ``\\Microsoft\\Windows\\TaskScheduler\\Task``).

    Returns:
        ``'System/Default'`` or ``'User-Created'``.
    """
    lower = task_path.lower()
    if lower.startswith("\\microsoft\\") or lower.startswith("\\windows\\"):
        return "System/Default"
    return "User-Created"


def collect_tasks(session) -> dict:
    """Enumerate all scheduled tasks on a remote server.

    Runs a PowerShell script via WinRM that collects every scheduled task,
    including its actions, triggers, run-as user, state, and last/next run
    times. Each task is classified as system/default or user-created.

    Args:
        session: A ``winrm.Session`` instance obtained from
            ``create_session``.

    Returns:
        A dictionary with three keys:

        - ``all_tasks``: list of all task dicts (sorted: user-created
          first, then system; alphabetical by name within each group).
        - ``user_tasks``: list of user-created task dicts.
        - ``system_tasks``: list of system/default task dicts.

        Each task dict contains: ``task_name``, ``task_path``,
        ``run_as_user``, ``triggers``, ``actions``, ``state``,
        ``last_run_time``, ``next_run_time``, ``classification``.

        Returns an empty dict if the operation fails on the server.
    """
    success, result = run_ps(session, _TASKS_SCRIPT)
    if not success:
        logger.error(
            "Failed to enumerate scheduled tasks: %s",
            result.get("error", "Unknown error"),
        )
        return {}

    if not result:
        return {}

    try:
        raw_tasks = json.loads(result)
    except json.JSONDecodeError as e:
        logger.error("Failed to parse scheduled task JSON: %s", e)
        return {}

    if isinstance(raw_tasks, dict):
        raw_tasks = [raw_tasks]
    elif not isinstance(raw_tasks, list):
        logger.warning(
            "Unexpected JSON structure from task enumeration: %s",
            type(raw_tasks),
        )
        return {}

    all_tasks = []
    for raw in raw_tasks:
        if not isinstance(raw, dict):
            continue

        task_name = raw.get("TaskName", "") or ""
        task_path = raw.get("TaskPath", "") or ""
        run_as_user = raw.get("RunAsUser", "") or ""
        triggers = raw.get("Triggers", "") or ""
        actions = raw.get("Actions", "") or ""
        state = raw.get("State", "") or ""
        last_run_time = raw.get("LastRunTime", "") or ""
        next_run_time = raw.get("NextRunTime", "") or ""
        classification = classify_task(task_path)

        all_tasks.append({
            "task_name": task_name,
            "task_path": task_path,
            "run_as_user": run_as_user,
            "triggers": triggers,
            "actions": actions,
            "state": state,
            "last_run_time": last_run_time,
            "next_run_time": next_run_time,
            "classification": classification,
        })

    all_tasks.sort(key=lambda t: (
        0 if t["classification"] == "User-Created" else 1,
        t["task_name"].lower(),
    ))

    user_tasks = [
        t for t in all_tasks if t["classification"] == "User-Created"
    ]
    system_tasks = [
        t for t in all_tasks if t["classification"] == "System/Default"
    ]

    return {
        "all_tasks": all_tasks,
        "user_tasks": user_tasks,
        "system_tasks": system_tasks,
    }


def build_summary_data(audit_data: dict) -> list[dict]:
    """Build per-server summary statistics.

    Args:
        audit_data: The full audit data structure.

    Returns:
        List of dicts suitable for ``write_summary``.
    """
    rows = []
    for s in audit_data["servers"]:
        users = s.get("users", [])
        groups = s.get("groups", [])
        tasks = s.get("tasks", {})
        rows.append({
            "server": s["server"],
            "total_users": len(users),
            "admins": sum(
                1 for u in users
                if u.get("privilege_level") == "Local Administrator"
            ),
            "privileged": sum(
                1 for u in users
                if u.get("privilege_level") == "Privileged User"
            ),
            "disabled": sum(
                1 for u in users
                if u.get("privilege_level") == "Disabled User"
            ),
            "builtin": sum(1 for u in users if u.get("is_builtin")),
            "groups": len(groups),
            "tasks": len(tasks.get("all_tasks", [])),
            "login_status": "Success" if s.get("login_success") else "Failed",
        })
    return rows


def build_login_status_data(audit_data: dict) -> list[dict]:
    """Build login status rows per server.

    Args:
        audit_data: The full audit data structure.

    Returns:
        List of dicts suitable for ``write_login_status``.
    """
    rows = []
    for s in audit_data["servers"]:
        rows.append({
            "server": s["server"],
            "success": s.get("login_success", False),
            "username_used": s.get("username_used", ""),
            "error": s.get("error") or "",
        })
    return rows


def flatten_users(audit_data: dict) -> list[dict]:
    """Flatten user data with server name for Excel output.

    Args:
        audit_data: The full audit data structure.

    Returns:
        List of dicts, one per user per server.
    """
    rows = []
    for s in audit_data["servers"]:
        for u in s.get("users", []):
            row = {"server": s["server"]}
            row.update(u)
            rows.append(row)
    return rows


def flatten_groups(audit_data: dict) -> list[dict]:
    """Flatten group-member data with server name for Excel output.

    Each group produces one row per member.

    Args:
        audit_data: The full audit data structure.

    Returns:
        List of dicts suitable for ``write_groups``.
    """
    rows = []
    for s in audit_data["servers"]:
        for g in s.get("groups", []):
            group_name = g.get("group_name", "")
            for m in g.get("members", []):
                rows.append({
                    "server": s["server"],
                    "group_name": group_name,
                    "member_name": m.get("name", ""),
                    "account_type": m.get("account_type", ""),
                    "enabled": m.get("enabled"),
                    "principal_source": m.get("principal_source", ""),
                })
    return rows


def flatten_tasks(audit_data: dict) -> list[dict]:
    """Flatten task data with server name for Excel output.

    Args:
        audit_data: The full audit data structure.

    Returns:
        List of dicts, one per task per server.
    """
    rows = []
    for s in audit_data["servers"]:
        for t in s.get("tasks", {}).get("all_tasks", []):
            row = {"server": s["server"]}
            row.update(t)
            rows.append(row)
    return rows


def flatten_errors(audit_data: dict) -> list[dict]:
    """Flatten per-operation error data with server name.

    Args:
        audit_data: The full audit data structure.

    Returns:
        List of dicts suitable for ``write_errors``.
    """
    rows = []
    for s in audit_data["servers"]:
        for e in s.get("errors", []):
            rows.append({
                "server": s["server"],
                "operation": e.get("operation", ""),
                "error": e.get("error", ""),
                "timestamp": e.get("timestamp", ""),
            })
    return rows


def generate_report(audit_data: dict, output_path: str) -> str:
    """Build and save the Excel audit report.

    Creates worksheets for Summary, Server Login Status, Local Users and
    Privileges, Local Groups and Members, Scheduled Tasks, and Errors
    (only if errors exist).  Saves the workbook to disk.

    Args:
        audit_data: The full audit data structure.
        output_path: Destination file path for the Excel file.

    Returns:
        The absolute path to the saved report.
    """
    wb = create_workbook()

    ws_summary = wb.create_sheet("Summary")
    write_summary(ws_summary, build_summary_data(audit_data))

    ws_login = wb.create_sheet("Server Login Status")
    write_login_status(ws_login, build_login_status_data(audit_data))

    ws_users = wb.create_sheet("Local Users and Privileges")
    write_users(ws_users, flatten_users(audit_data))

    ws_groups = wb.create_sheet("Local Groups and Members")
    write_groups(ws_groups, flatten_groups(audit_data))

    ws_tasks = wb.create_sheet("Scheduled Tasks")
    write_tasks(ws_tasks, flatten_tasks(audit_data))

    errors_data = flatten_errors(audit_data)
    if errors_data:
        ws_errors = wb.create_sheet("Errors")
        write_errors(ws_errors, errors_data)

    save_report(wb, output_path)
    return os.path.abspath(output_path)


def _setup_file_logging(log_path: str) -> None:
    log_dir = os.path.dirname(log_path)
    if log_dir:
        os.makedirs(log_dir, exist_ok=True)
    handler = logging.FileHandler(log_path, encoding="utf-8", mode="w")
    handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    )
    logging.getLogger().addHandler(handler)
    logging.getLogger().setLevel(logging.INFO)
    logging.getLogger("openpyxl").setLevel(logging.WARNING)
    logging.getLogger("urllib3").setLevel(logging.WARNING)


def main():
    """Entry point for the read-only audit."""
    parser = argparse.ArgumentParser(
        description="Windows Server Audit - Read-Only"
    )
    parser.add_argument(
        "--servers",
        default=os.path.join("config", "servers.txt"),
        help="Path to server list (default: config/servers.txt)",
    )
    parser.add_argument(
        "--key",
        default=os.path.join("config", "secret.key"),
        help="Path to encryption key (default: config/secret.key)",
    )
    parser.add_argument(
        "--passwords",
        default=os.path.join("config", "passwords.enc"),
        help="Path to encrypted passwords (default: config/passwords.enc)",
    )
    parser.add_argument(
        "--output",
        default="output",
        help="Output directory for reports (default: output/)",
    )
    args = parser.parse_args()

    os.makedirs(args.output, exist_ok=True)
    ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
    _setup_file_logging(os.path.join(args.output, f"Audit_{ts_file}.log"))

    servers = load_servers(args.servers)
    if not servers:
        print("No servers found in server list.", file=sys.stderr)
        logger.error("No servers found in server list.")
        sys.exit(1)

    if not os.path.exists(args.key):
        print(
            f"Error: Encryption key not found at {args.key}",
            file=sys.stderr,
        )
        logger.error("Encryption key not found at %s", args.key)
        sys.exit(1)

    if not os.path.exists(args.passwords):
        print(
            f"Error: Encrypted passwords not found at {args.passwords}",
            file=sys.stderr,
        )
        logger.error("Encrypted passwords not found at %s", args.passwords)
        sys.exit(1)

    passwords = read_encrypted_file(args.key, args.passwords)
    usernames = ["administrator", "admin"]

    audit_data: dict = {"servers": []}

    for server in servers:
        server_entry: dict = {
            "server": server,
            "login_success": False,
            "username_used": "",
            "error": None,
            "users": [],
            "groups": [],
            "tasks": {},
            "errors": [],
        }

        logger.info("[%s] Connecting...", server)
        sys.stdout.write(f"[{server}] Connecting... ")
        sys.stdout.flush()
        session, result = create_session(server, usernames, passwords)

        if session is not None:
            server_entry["login_success"] = True
            server_entry["username_used"] = result
            logger.info("[%s] OK (%s)", server, result)
            print(f"OK ({result})")

            sys.stdout.write(f"[{server}] Collecting users... ")
            sys.stdout.flush()
            try:
                users = collect_users(session)
                server_entry["users"] = users
                logger.info("[%s] %d users found", server, len(users))
                print(f"{len(users)} users found")
            except Exception as e:
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                server_entry["errors"].append({
                    "operation": "collect_users",
                    "error": str(e),
                    "timestamp": ts,
                })
                logger.error("[%s] collect_users ERROR - %s", server, e)
                print(f"ERROR - {e}")

            sys.stdout.write(f"[{server}] Collecting groups... ")
            sys.stdout.flush()
            try:
                groups = collect_groups(session)
                server_entry["groups"] = groups
                logger.info("[%s] %d groups found", server, len(groups))
                print(f"{len(groups)} groups found")
            except Exception as e:
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                server_entry["errors"].append({
                    "operation": "collect_groups",
                    "error": str(e),
                    "timestamp": ts,
                })
                logger.error("[%s] collect_groups ERROR - %s", server, e)
                print(f"ERROR - {e}")

            sys.stdout.write(f"[{server}] Collecting tasks... ")
            sys.stdout.flush()
            try:
                tasks = collect_tasks(session)
                server_entry["tasks"] = tasks
                all_count = len(tasks.get("all_tasks", []))
                user_count = len(tasks.get("user_tasks", []))
                logger.info("[%s] %d tasks (%d user-created)", server, all_count, user_count)
                print(f"{all_count} tasks ({user_count} user-created)")
            except Exception as e:
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                server_entry["errors"].append({
                    "operation": "collect_tasks",
                    "error": str(e),
                    "timestamp": ts,
                })
                logger.error("[%s] collect_tasks ERROR - %s", server, e)
                print(f"ERROR - {e}")

            close_session(session)
        else:
            server_entry["error"] = result
            logger.error("[%s] FAILED - %s", server, result)
            print(f"FAILED - {result}")
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            server_entry["errors"].append({
                "operation": "connect",
                "error": result,
                "timestamp": ts,
            })

        audit_data["servers"].append(server_entry)

    print("---")

    filename = f"Audit_Report_{ts_file}.xlsx"
    output_path = os.path.join(args.output, filename)

    report_path = generate_report(audit_data, output_path)
    logger.info("Audit complete. Report saved to %s", report_path)
    print(f"Audit complete. Report saved to {report_path}")


if __name__ == "__main__":
    main()
