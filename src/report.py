"""Formatted Excel workbook generation for Windows Server Audit and
password change reports.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(
    start_color="2F5496", end_color="2F5496", fill_type="solid"
)
HEADER_ALIGNMENT = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_workbook() -> Workbook:
    """Create a new Excel workbook with the default sheet removed.

    Returns:
        An empty ``Workbook`` instance ready for worksheets to be added.
    """
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _add_header_style(ws, headers: list[str]) -> None:
    """Write formatted header row and configure auto-filter and freeze.

    Writes headers to row 1, applies font, fill, alignment, and border
    styling, adds an auto-filter over the header range, and freezes the
    top row for scrolling.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        headers: Column header strings to write in order.
    """
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.freeze_panes = "A2"


def _auto_width(ws) -> None:
    """Adjust column widths based on content length, capped at 60.

    Iterates over every column in the worksheet and sets the column
    width to the length of the longest value (header or cell) plus a
    padding of 2, with a maximum width of 60 characters.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
    """
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except (ValueError, TypeError):
                pass
        adjusted = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = adjusted


def write_summary(ws, data: list[dict]) -> None:
    """Write the Summary worksheet with per-server overview counts.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``total_users``,
            ``admins``, ``privileged``, ``disabled``, ``builtin``,
            ``groups``, ``tasks``, ``login_status``.
    """
    headers = [
        "Server",
        "Total Users",
        "Admins",
        "Privileged",
        "Disabled",
        "Built-in",
        "Groups",
        "Tasks",
        "Login Status",
    ]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("total_users", 0))
        ws.cell(row=row_idx, column=3, value=entry.get("admins", 0))
        ws.cell(row=row_idx, column=4, value=entry.get("privileged", 0))
        ws.cell(row=row_idx, column=5, value=entry.get("disabled", 0))
        ws.cell(row=row_idx, column=6, value=entry.get("builtin", 0))
        ws.cell(row=row_idx, column=7, value=entry.get("groups", 0))
        ws.cell(row=row_idx, column=8, value=entry.get("tasks", 0))
        ws.cell(row=row_idx, column=9, value=entry.get("login_status", ""))
    _auto_width(ws)


def write_login_status(ws, data: list[dict]) -> None:
    """Write the Server Login Status worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``success``,
            ``username_used``, ``error``.
    """
    headers = ["Server", "Success", "Username Used", "Error"]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("success", False))
        ws.cell(
            row=row_idx, column=3, value=entry.get("username_used", "")
        )
        ws.cell(row=row_idx, column=4, value=entry.get("error", ""))
    _auto_width(ws)


def write_users(ws, data: list[dict]) -> None:
    """Write the Local Users and Privileges worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``name``, ``fullname``,
            ``enabled``, ``privilege_level``, ``privileged_groups``,
            ``principal_source``, ``is_builtin``.
    """
    headers = [
        "Server",
        "Name",
        "Full Name",
        "Enabled",
        "Privilege Level",
        "Privileged Groups",
        "Principal Source",
        "Is Built-in",
    ]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("name", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("fullname", ""))
        ws.cell(row=row_idx, column=4, value=entry.get("enabled", False))
        ws.cell(
            row=row_idx,
            column=5,
            value=entry.get("privilege_level", ""),
        )
        ws.cell(
            row=row_idx,
            column=6,
            value=entry.get("privileged_groups", ""),
        )
        ws.cell(
            row=row_idx,
            column=7,
            value=entry.get("principal_source", ""),
        )
        ws.cell(row=row_idx, column=8, value=entry.get("is_builtin", False))
    _auto_width(ws)


def write_groups(ws, data: list[dict]) -> None:
    """Write the Local Groups and Members worksheet.

    Each dict represents one group-member row (data must be flattened
    before calling this function).

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``group_name``,
            ``member_name``, ``account_type``, ``enabled``,
            ``principal_source``.
    """
    headers = [
        "Server",
        "Group Name",
        "Member Name",
        "Account Type",
        "Enabled",
        "Principal Source",
    ]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("group_name", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("member_name", ""))
        ws.cell(
            row=row_idx, column=4, value=entry.get("account_type", "")
        )
        ws.cell(row=row_idx, column=5, value=entry.get("enabled"))
        ws.cell(
            row=row_idx,
            column=6,
            value=entry.get("principal_source", ""),
        )
    _auto_width(ws)


def write_tasks(ws, data: list[dict]) -> None:
    """Write the Scheduled Tasks worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``task_name``,
            ``task_path``, ``run_as_user``, ``triggers``, ``actions``,
            ``state``, ``last_run_time``, ``next_run_time``,
            ``classification``.
    """
    headers = [
        "Server",
        "Task Name",
        "Path",
        "Run As User",
        "Triggers",
        "Actions",
        "State",
        "Last Run",
        "Next Run",
        "Classification",
    ]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("task_name", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("task_path", ""))
        ws.cell(
            row=row_idx, column=4, value=entry.get("run_as_user", "")
        )
        ws.cell(row=row_idx, column=5, value=entry.get("triggers", ""))
        ws.cell(row=row_idx, column=6, value=entry.get("actions", ""))
        ws.cell(row=row_idx, column=7, value=entry.get("state", ""))
        ws.cell(
            row=row_idx, column=8, value=entry.get("last_run_time", "")
        )
        ws.cell(
            row=row_idx, column=9, value=entry.get("next_run_time", "")
        )
        ws.cell(
            row=row_idx,
            column=10,
            value=entry.get("classification", ""),
        )
    _auto_width(ws)


def write_password_status(ws, data: list[dict]) -> None:
    """Write the Password Change Status worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``old_connection_status``,
            ``new_password_set``, ``error``.
    """
    headers = ["Server", "Old Connection Status", "New Password Set", "Error"]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(
            row=row_idx,
            column=2,
            value=entry.get("old_connection_status", ""),
        )
        ws.cell(
            row=row_idx,
            column=3,
            value=entry.get("new_password_set", False),
        )
        ws.cell(row=row_idx, column=4, value=entry.get("error", ""))
    _auto_width(ws)


def write_errors(ws, data: list[dict]) -> None:
    """Write the Errors worksheet.

    Args:
        ws: An openpyxl ``Worksheet`` instance.
        data: List of dicts with keys ``server``, ``operation``,
            ``error``, ``timestamp``.
    """
    headers = ["Server", "Operation", "Error", "Timestamp"]
    _add_header_style(ws, headers)
    for row_idx, entry in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("server", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("operation", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("error", ""))
        ws.cell(row=row_idx, column=4, value=entry.get("timestamp", ""))
    _auto_width(ws)


def save_report(workbook, path: str) -> None:
    """Save a workbook to disk.

    Args:
        workbook: An openpyxl ``Workbook`` instance to save.
        path: Destination file path.
    """
    workbook.save(path)
